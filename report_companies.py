#!/usr/bin/env python
# coding: utf-8
from configs import *
import os
import re
import pandas as pd
import owncloud
import numpy as np
import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
from openpyxl.utils import get_column_letter
import xlwings as xlw
import warnings
from datetime import datetime, date, timedelta
from time import sleep
import psycopg2
import urllib3

#Игнорируем всплывающие предупреждения
warnings.filterwarnings('ignore')

# OwnCloud
REMOTE_DIR = '<replaced>'

# Меняем рабочую папку
LOCAL_PATH = r'<replaced>'
os.chdir(LOCAL_PATH)

SQL_MON = """select 
                (fm."Rep_Date" + 1)::date as "Rep_Date"
                ,fm."<replaced>" 
                ,sum(fm."<replaced>") as "<replaced>"
                ,sum(fm."<replaced>") as "<replaced>"
                ,sum(fm."<replaced>") as "<replaced>"
                ,sum(fm."<replaced>") as "<replaced>"
                ,sum(fm."<replaced>") as "<replaced>"
                ,sum(fm."<replaced>") as "<replaced>"
            from "<replaced>" fm 
            where fm."Rep_Date" >= '2022-03-24'
            and date_part('dow', fm."Rep_Date") = 4
            group by 1, 2"""


# SEND FUNC 
def send_report(remote_dir, file, user, password):
    # Auth OwnCloud
    urllib3.disable_warnings()
    oc = owncloud.Client('<replaced>', verify_certs=False)
    oc.login(user, password)
    print('Auth OwnCloud ok.')
    
    # Send report
    path = remote_dir + "/" +  file
    oc.put_file(path, file)


#Вытаскиваем файл с данными о компаниях РВР с фтп-сфтп сервера
def get_file_sftp_ftp(connection, host, port, user, password, path, filename,
                      private_key_path=None, private_key_pass=None):
    
    if connection == 'sftp':
        
        # Establish connection
        import pysftp
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None
        sftp = pysftp.Connection(host, username=user, private_key=private_key_path,
                                 private_key_pass=private_key_pass, cnopts=cnopts)
        sftp.chdir(path)
        print ("SFTP-connection succesfully established.")
        
        # Get all files matching filename
        files = [(file.filename, file.st_mtime) for file in sftp.listdir_attr() if file.filename.startswith(filename)]
        files = [(file[0], datetime.fromtimestamp(file[1])) for file in files]
        filename = max(files, key=lambda x: x[1])[0]
        sftp.get(filename,  filename)
        sftp.close()
        
    else:
        #establish connection
        from ftplib import FTP
        ftp = FTP(host, user, password)
        ftp.cwd(path)
        ftp.encoding = 'utf-8'
        print ("FTP-connection succesfully established.")
        
        # Get all files matching filename
        files = [(file[0], file[1]['modify']) for file in ftp.mlsd(path) if file[0].startswith(filename)]
        filename = max(files, key=lambda x: x[1])[0]

        with open(filename, 'wb') as local_file:
            ftp.retrbinary(f'RETR {filename}', local_file.write)

    return filename


def connect_and_execute(sql, cols=None):
    # Connect
    conn = psycopg2.connect(dbname='<replaced>', user='<replaced>',
                            password='<replaced>', host='<replaced>', port='<replaced>')
    cursor = conn.cursor()
    
    # Execute
    cursor.execute(sql)
    
    # Collect
    df = pd.DataFrame(cursor.fetchall(), columns=cols)
    return df

def get_pfr():
    # GET PFR
    host = "<replaced>"
    port = 21
    user = "<replaced>"
    password = "<replaced>"
    path = '<replaced>'
    shablon = '<replaced>'
    file = get_file_sftp_ftp('ftp', host, port, user, password, path, shablon)

    return file

def process_pfr(df, cols_company):
    # GET all columns with dates (they contain number of employees)
    cols_PFR = df.columns.to_list()
    cols_dates = [col for col in cols_PFR if type(col)==datetime]

    # Get columns with main data
    cols_new = cols_company + cols_dates
    df = df[cols_new]

    # Multiindex columns
    cols_new = [(col, '') for col in cols_company] + [('<replaced>', col.date()) for col in cols_dates]
    df.columns = pd.MultiIndex.from_tuples(cols_new)
    df = df.drop_duplicates()
    df = df.loc[df[('<replaced>', '')].isna()==False]
    
    return df

def process_rvr(df, cols):
    # MAKE RVR pivot
    index = ['<replaced>']
    values = cols[2:]
    columns = ['Rep_Date']
    aggfunc = {k:'sum' for k in values}

    df = pd.pivot_table(df, index=index, values=values, columns=columns, aggfunc=aggfunc)
    df.reset_index(inplace=True)

    return df


def process_all(df_PFR, df_RVR, cols_company):
    # Preprocessing for merging
    df_PFR[('<replaced>', '')] = df_PFR[('<replaced>', '')].astype('int64')
    df_RVR[('<replaced>', '')] = df_RVR[('<replaced>', '')].astype('int64')

    # Merging
    df = df_PFR.merge(df_RVR, how='left', on='<replaced>')
    df.drop('<replaced>', axis=1, inplace=True)
    
    # Process data columns
    cols_all = df.columns.to_list()
    df[cols_all[len(cols_company):]] = df[cols_all[len(cols_company):]].fillna(0).astype('int64')
    
    # Calculate diff
    num_PFR = '<replaced>'
    num_RVR = '<replaced>'
    num_DIFF = '<replaced>'
    num_NEW = '<replaced>'
    dates_PFR = [col[1] for col in cols_all if col[0]==num_PFR]
    for date in dates_PFR:
        df[(num_DIFF, date)] = df[(num_PFR, date)] - df[(num_RVR, date)]

    # Sort columns for report
    for date in dates_PFR:
        date_string = date.strftime('%d.%m.%Y')
        df[(num_NEW, f'{date_string} <replaced>')] = df[(num_PFR, date)] 
        df[(num_NEW, f'{date_string} <replaced>')] = df[(num_RVR, date)]
        df[(num_NEW, f'{date_string} <replaced>')] = df[(num_DIFF, date)]

    # Order dates
    cols_data = ['<replaced>', '<replaced>', '<replaced>',
                 '<replaced>', '<replaced>', '<replaced>']

    # Right order columns
    cols_new = cols_company + cols_data
    df = df[cols_new]
    
    # RENAME
    cols_rus = ['<replaced>', 
                '<replaced>',
                '<replaced>', 
                '<replaced>', 
                '<replaced>',
                '<replaced>']
    df = df.rename(columns=dict(zip(cols_data, cols_rus)))

    return df

def copy_sheet_PFR(file_PFR, file_report):
    #Копируем из файла ПФР первый лист в новый отчет
    app = xlw.App(visible=False)
    wb = xlw.Book(file_PFR)
    sheet = wb.sheets[0]

    new_wb = xlw.Book(file_report)
    sheet.api.Copy(Before=new_wb.sheets[0].api)
    new_wb.sheets[0].name = '<replaced>'
    new_wb.sheets[1].name = '<replaced>'
    
    new_wb.save(file_report)
    wb.close()
    new_wb.close()


def style_report(file):
    # LOAD REPORT
    wb = xl.load_workbook(file) # GET WB
    sheet = wb['<replaced>'] # GET SHEET PFR
    sheet.sheet_view.zoomScale = 70 # SET ZOOM
    sheet.sheet_view.topLeftCell = 'A1'
    sheet.freeze_panes = 'K2' 
    
    # GET SHEET PFR + RVR
    sheet = wb['<replaced>']
    
    # DECALRE STYLES
    no_border = Border(left=Side(style=None), right=Side(style=None), 
                       top=Side(style=None), bottom=Side(style=None))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    redFill = PatternFill(fill_type='solid', start_color='00FFDAB9', end_color='00FFDAB9')
    greenFill = PatternFill(fill_type='solid', start_color='008FBC8F', end_color='008FBC8F')
    paleGreenFill = PatternFill(fill_type='solid', start_color='00B4EEB4', end_color='00B4EEB4')
    
    # SET SHEET TITLE
    sheet.delete_cols(1)
    sheet.insert_cols(1)
    
    # SET CELL VALUES
    # переносим названия колонок с 1ой строки на 2ую
    for col in range(2, 12):
        val = sheet.cell(row=1, column=col).value
        sheet.cell(row=2, column=col, value=val)
        sheet.cell(row=1, column=col).value = None
        sheet.cell(row=1, column=col).border = no_border
    
    # ячейки с данными компаний
    for row in sheet['B2':f'K{sheet.max_row}']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = Font(name='Times New Roman', size=11)
            
    column_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    column_width = [15 for i in column_letters]
    for i in ['E', 'K', 'J']:
        column_width[column_letters.index(i)] = 20

    for col, width in zip(column_letters, column_width):
        sheet.column_dimensions[col].width = width
        
            
    # SET BORDER AND ALIGNMENT
    # 1ая строка
    for col_num in range(11, sheet.max_column+1):
        letter = get_column_letter(col_num)
        sheet.column_dimensions[letter].width = 12
        
    letter = get_column_letter(sheet.max_column)
    for row in sheet['L1':f'{letter}1']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = greenFill
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            
    # 2ая строка
    for row in sheet['B2':f'{letter}2']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = redFill
            cell.font = Font(name='Times New Roman', size=11, bold=True)
    sheet.row_dimensions[2].height = 42
    
            
    # 3ья строка
    for row in sheet['B3':f'{letter}3']:
        for ind, cell in enumerate(row, start=1):
            cell.value = ind
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = paleGreenFill
            cell.font = Font(name='Times New Roman', size=11, bold=True)
    
            
    # SHEET STYLE
    sheet.sheet_view.topLeftCell = 'A1'
    sheet.freeze_panes = 'L4'
    sheet.sheet_view.zoomScale = 70
    
    # SAVE
    wb.save(file)
    wb.close()

def main():
    # GET PFR FILE WITH DATA
    file_PFR = get_pfr()
    df_PFR = pd.read_excel(file_PFR, dtype='object')
    print('<replaced>:', file_PFR)
    
    # GET RVR DATA FROM DB
    cols_RVR = ['Rep_Date', '<replaced>',
                '<replaced>', '<replaced>', '<replaced>',
                '<replaced>', '<replaced>', '<replaced>']
    df_RVR = connect_and_execute(SQL_MON, cols=cols_RVR)
    df_RVR.to_csv('<replaced>.csv', sep=';', encoding='utf-8', index=False)
    
    # PROCESS PFR
    cols_company = ['<replaced>']
    df_PFR = process_pfr(df_PFR, cols_company)
    print("First part processed.")
    
    # PROCESS RVR
    df_RVR = process_rvr(df_RVR, cols_RVR)
    print("Second part processed.")
    
    # MERGE PFR AND RVR
    df_REPORT = process_all(df_PFR, df_RVR, cols_company)
    print("Data unioned.")
    
    # SAVE
    date_report = date.today()
    file_report = f"<replaced> ({date_report.strftime('%Y-%m-%d')}).xlsx"
    df_REPORT.to_excel(file_report)
    print("Report created with name:", file_report)
    
    # COPY SHEET
    copy_sheet_PFR(file_PFR, file_report)
    print('Sheet <replaced> copied to new report.')
    
    # STYLE
    style_report(file_report)
    print('Styling finished')
    
    # SEND
    send_report(REMOTE_DIR, file_report, USER_OC, PASS_OC)
    print('Report sent.')


if __name__ == '__main__':
       main()
